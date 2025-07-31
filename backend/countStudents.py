import openpyxl
import sys
import json
import os
def main():
    
    sys.stdout.flush()
    try : 
        data_json = sys.stdin.read()
        data = json.loads(data_json)

        details = data['details']
        rooms = data.get('rooms', [])
        sem = details[0].get("sem")
        slots = {item.get("slot") for item in details}
        total_students = 0

        for slot in slots:
            branches = [item.get("branch") for item in details if item.get("slot") == slot]
            for branch in branches:
                        file_path = f'./updatedExcels/Sem{sem}_{branch}.xlsx'
                        
                        try:
                            wb = openpyxl.load_workbook(file_path)
                            sheetnames_lc_map = {s.lower(): s for s in wb.sheetnames} 
                            slot_lc = slot.lower()
                            if slot_lc in sheetnames_lc_map:
                                ws_main = wb[sheetnames_lc_map[slot_lc]]
                                total_students += ws_main.max_row - 1
                            
                        except Exception as e:
                            print(f"Error opening {file_path}: {e}", file=sys.stderr)
                            sys.exit(1)
        assigned_students = 0
        rooms_used = []
        for room in rooms:
            if assigned_students >= total_students:
                break
            room_name = room.get("room_no")
            capacity = room.get("capacity", 0)
            assigned_students += capacity
            rooms_used.append({
                "room": room_name,
                "capacity": capacity
            })
        if assigned_students < total_students:
                    print(f"Error: Not enough capacity for {total_students} students.", file=sys.stderr)
                    sys.exit(1)
        output = {
        "total_students": total_students,
        "rooms_used": rooms_used
        }
        print(json.dumps(total_students))

          
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()

