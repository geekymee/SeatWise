import sys
import json
import openpyxl
from openpyxl import Workbook
from pathlib import Path

def load_input_files():
    data = json.load(sys.stdin)
    return data
   
def get_branch_names(ws):
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_map = {col_name.strip(): idx for idx, col_name in enumerate(header) if col_name}
    
    if 'Branch' not in col_map:
        raise ValueError("Missing 'Branch' column in the sheet.")
    
    branch_col_idx = col_map['Branch'] + 1 

    branches = set()
    for row in ws.iter_rows(min_row=2, min_col=branch_col_idx, max_col=branch_col_idx, values_only=True):
        val = row[0]
        if val and val != 'Branch Name':
            branches.add(val.strip())
    
    return list(branches)


def create_main_sheet(ws, branch_name):
    main_data = []
   
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_map = {str(col_name).strip(): idx for idx, col_name in enumerate(header) if col_name}

    required_cols = ['Name', 'RegNo', 'Branch', 'Slot', 'Subject Code']
 
    for col in required_cols:
        if col not in col_map:
            raise ValueError(f"Missing required column: '{col}' in the sheet.")

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[col_map['Name']]
        regno = row[col_map['RegNo']]
        branch = row[col_map['Branch']]

        if all([branch, branch_name, name, regno]) and branch.strip().lower() == branch_name.strip().lower():
            subcode_val = row[col_map['Subject Code']]
            subcode = subcode_val.strip() if subcode_val else ""
            slot = row[col_map['Slot']]
            main_data.append((name, regno, branch_name, slot, subcode))
    return main_data
   
def write_main_sheet(ws, data, headers):
    for c, header in enumerate(headers, 1):
        ws.cell(row=1, column=c).value = header
    for r, row in enumerate(data, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c).value = val

def sort_and_write_temp_sheet(wb, main_data):
    temp_sheet = wb.create_sheet('Temp')
    sorted_data = sorted(set(main_data), key=lambda x: x[1])  
    for row in sorted_data:
        temp_sheet.append(row)
    return temp_sheet

def create_slot_sheets(wb, temp_ws, codeno):
    slot_col = 4
    regno_col = 2
    subcode_col = 5
    slot_data = {}
    raw_slots = []
    for row in temp_ws.iter_rows(min_row=1, values_only=True):
        raw_value = row[slot_col-1]
        if raw_value:
            normalized = str(raw_value).strip().capitalize()
            raw_slots.append(normalized)
    
    slots = sorted(set(raw_slots))

    for slot in slots:
        
        sheet = wb.create_sheet(slot)
        r = 1

        
        for row in temp_ws.iter_rows(min_row=1, values_only=True):
            current_slot = str(row[slot_col - 1]).strip().capitalize()
            if current_slot == slot:
                for c, val in enumerate((row[0], row[regno_col - 1], current_slot, row[subcode_col - 1]), 1):
                    sheet.cell(row=r, column=c).value = val
                r += 1
        
       
        slot_data[codeno + slot] = r - 1  

    return slot_data

def process_file(filename):
    wb = openpyxl.load_workbook(f'./uploadedExcels/{filename}')
    ws = wb[wb.sheetnames[0]]
    sem_prefix = filename.split('_')[0]

    branches = get_branch_names(ws)
    all_slot_data = []

    for branch in branches:
        branch_wb = Workbook()
        main_ws = branch_wb.active
        main_ws.title = "Main"

        main_data = create_main_sheet(ws, branch)
        if not main_data:
            print(f"[SKIPPED] No data found for branch: {branch}")
            continue

        write_main_sheet(main_ws, main_data, headers=["Name", "RegNo", "Branch", "Slot", "Subject Code"])

        regno_example = str(main_data[0][1])
        codeno = regno_example[4:8] if len(regno_example) >= 8 else "XXXX"

        filepath = Path(f'./updatedExcels/{sem_prefix}_{branch}.xlsx')
        print(f"filepath: {filepath}")
        
        temp_ws = sort_and_write_temp_sheet(branch_wb, main_data)
        
        slot_data = create_slot_sheets(branch_wb, temp_ws, codeno)
        branch_wb.save(filepath)
        all_slot_data.append(slot_data)

    return all_slot_data

def main():
    # if isinstance(data, dict) and data.get('mode') == 'subjectCodes':
    #     filename = data.get('filename')
    #     if not filename:
    #         print(json.dumps({"error": "Filename is required"}))
    #         sys.exit(1)
    #     try:
    #         wb = openpyxl.load_workbook(f'./uploadedExcels/{filename}')
    #         ws = wb[wb.sheetnames[0]]
    #         subject_codes = get_subjectCode(ws)
    #         print(json.dumps({"subjectCodes": subject_codes}))
    #     except Exception as e:
    #         print(json.dumps({"error": str(e)}))
    #         sys.exit(1)
    # else:
    #     results = []
        
    filenames = load_input_files()
    for file in filenames:
        result = process_file(file)
        print(f"Processed {file}: {result}")
        result = process_file(file)
            # results.append({"filename": file, "slot_data": result})
        # print(json.dumps({"results": results}))

if __name__ == "__main__":
    main()
