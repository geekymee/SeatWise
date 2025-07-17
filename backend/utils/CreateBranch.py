import sys
import json
import openpyxl
from openpyxl import Workbook
from pathlib import Path

def load_input_files():
    data = json.load(sys.stdin)
    return data

def get_branch_names(ws):
    branches = {cell.value for cell in ws['D'] if cell.value not in (None, 'Branch Name')}
    return list(branches)

def create_main_sheet(ws, branch_name):
    main_data = []

   
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_map = {col_name.strip(): idx for idx, col_name in enumerate(header)}

  
    required_cols = ['Name', 'RegNo', 'Branch', 'Slot', 'Subject Code']

 
    for col in required_cols:
        if col not in col_map:
            raise ValueError(f"Missing required column: '{col}' in the sheet.")

   
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[col_map['Name']]
        regno = row[col_map['RegNo']]
        branch = row[col_map['Branch']]

        if branch == branch_name and name and regno:
            subcode_val = row[col_map['Subject Code']]
            subcode = subcode_val.strip() if subcode_val else ""
            slot = row[col_map['Slot']]
            main_data.append((name, regno, branch_name, slot, subcode))

    return main_data


def write_main_sheet(ws, data):
    for r, row in enumerate(data, 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c).value = val

def sort_and_write_temp_sheet(wb, main_data):
    temp_sheet = wb.create_sheet('Temp')
    sorted_data = sorted(set(main_data), key=lambda x: x[1])  
    for row in sorted_data:
        temp_sheet.append(row)
    return temp_sheet

def create_slot_sheets(wb, temp_ws, codeno):
    slot_col = 4
    regno_col = 2
    subcode_col = 5
    slot_data = {}

    slots = sorted({cell.value for cell in temp_ws.iter_rows(min_row=1, max_col=slot_col, values_only=True) if cell[slot_col - 1]})
    
    for slot in slots:
        regular = wb.create_sheet(slot)
        supply = None
        regnos = [row[regno_col - 1] for row in temp_ws.iter_rows(min_row=1, values_only=True) if row[slot_col - 1] == slot]
        years = sorted({
            int(reg[2:3]) for reg in regnos
            if isinstance(reg, str) and len(reg) >= 5 and reg[2:3].isdigit()
        })
        if len(years) > 1:
            supply = wb.create_sheet(slot + "_supply")

        r_reg, r_sup = 1, 1
        for row in temp_ws.iter_rows(min_row=1, values_only=True):
            if row[slot_col - 1] == slot:
                reg = row[regno_col - 1]
                target_ws = regular if reg[3:5] == years[-1] else supply
                if target_ws:
                    for c, val in enumerate((row[0], reg, row[slot_col - 1], row[subcode_col - 1]), 1):
                        target_ws.cell(row=(r_reg if target_ws == regular else r_sup), column=c).value = val
                    if target_ws == regular:
                        r_reg += 1
                    else:
                        r_sup += 1

        slot_data[codeno + slot] = r_reg - 1
    return slot_data

def process_file(filename):
    wb = openpyxl.load_workbook(f'./uploadedExcels/{filename}')
    ws = wb[wb.sheetnames[0]]
    sem_prefix = filename[:2]

    branches = get_branch_names(ws)
    all_slot_data = []

    for branch in branches:
        branch_wb = Workbook()
        main_ws = branch_wb.active
        main_ws.title = "Main"

        main_data = create_main_sheet(ws, branch)
        if not main_data:
            print(f"[SKIPPED] No data found for branch: {branch}")
            continue

        write_main_sheet(main_ws, main_data, headers=["Name", "RegNo", "Branch", "Slot", "Subject Code"])

        regno_example = main_data[0][1]
        codeno = regno_example[4:8] if len(regno_example) >= 8 else "XXXX"

        filepath = Path(f'./updatedExcels/{sem_prefix}_{codeno}_{branch}.xlsx')
        temp_ws = sort_and_write_temp_sheet(branch_wb, main_data)
        slot_data = create_slot_sheets(branch_wb, temp_ws, codeno)

        branch_wb.save(filepath)
        all_slot_data.append(slot_data)

    return all_slot_data

def main():
    filenames = load_input_files()
    for file in filenames:
        result = process_file(file)
        print(f"Processed {file}: {result}")

if __name__ == "__main__":
    main()
