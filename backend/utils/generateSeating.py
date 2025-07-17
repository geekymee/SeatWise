import openpyxl
import math
from openpyxl import Workbook
from openpyxl.styles import Alignment 
import sys
import json

def main():
    try:
        data_json = sys.stdin.read() 
        data = json.loads(data_json)

        details = data['details']
        rooms = data.get('rooms', [])
        sem = details[0].get("sem")

        slots = {item.get("slot") for item in details}
        for slot in slots:
            branches = [item.get("branch") for item in details if item.get("slot") == slot]
            students = []
            for branch in branches:
                file_path = f'./updatedExcels/S{sem}_{branch}.xlsx'
                try:
                    wb = openpyxl.load_workbook(file_path)
                    if slot in wb.sheetnames:
                        ws_main = wb[slot]
                        for row in ws_main.iter_rows(min_row=1, max_row=ws_main.max_row, values_only=True):
                            if any(row):
                                students.append(list(row) + [branch])
                    supply_sheet = f"{slot}_supply"
                    if supply_sheet in wb.sheetnames:
                        ws_supply = wb[supply_sheet]
                        for row in ws_supply.iter_rows(min_row=1, max_row=ws_supply.max_row, values_only=True):
                            if any(row):
                                students.append(list(row) + [branch, "SUPPLY"])
                except Exception as e:
                    print(f"Error opening {file_path}: {e}", file=sys.stderr)
                    sys.exit(1)
            total_students = len(students)
            assigned = 0

            room_data = []
            wb_output = openpyxl.Workbook()
            del wb_output["Sheet"]
        
            for room in rooms:
                if assigned >= total_students:
                    break

                room_name = room["room"]
                capacity = room["capacity"]
                ws = wb_output.create_sheet(title=room_name)
                ws.append(["Branch", "Seat No.", "Sub. code", "Reg. No", "From Branch", "Type"])
                rowA = rowB = 2
                half = capacity // 2
            
            for i in range(capacity):
                if assigned >= total_students:
                    break
                student = students[assigned]
                target_col = 1 if i < half else 4  
                target_row = rowA if i < half else rowB
                for c, value in enumerate(student):
                    ws.cell(row=target_row, column=target_col + c, value=value)
                if i < half:
                    rowA += 1
                else:
                    rowB += 1
                assigned += 1
            
            room_data.append({
                "room": room_name,
                "capacity": capacity,
                "assigned": rowA + rowB - 2  # approx
            })
                
        if assigned < total_students:
                print(f"Error: Not enough capacity for {total_students} students.", file=sys.stderr)
                sys.exit(1)

        
        output_filename = f"./output/Seating_Sem{sem}_Slot_{slot}.xlsx"
        wb_output.save(output_filename)   

            
            

            

        for room_sheet in wb_output.sheetnames:
            ws = wb_output[room_sheet]
            # Skip if no data
            if ws.max_row < 3:
                continue
            i = 0
            start = 2
            flag = 0
            br = ws.cell(row=2, column=4).value[5:7] if ws.cell(row=2, column=4).value else ""
            for p in range(2, ws.max_row + 1):
                current_cell = ws.cell(row=p, column=4).value
                if not current_cell:
                    continue
                current_br = current_cell[5:7]
                if (br != current_br) or (p == ws.max_row):
                    end_row = p - 1 if br != current_br else p
                    s = f"A{start}:A{end_row}"
                    ws.merge_cells(s)
                    cell = ws.cell(row=start, column=1)
                    cell.value = br
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    br = current_br
                    start = p
                i += 1
        wb_output.save(output_filename)
        
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()