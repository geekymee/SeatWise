import openpyxl
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment 
import sys
import json

def main():
    try:
        data_json = sys.stdin.read() 
        data = json.loads(data_json)
        
        
        details = data['details']
        rooms = data.get('rooms', [])
        sem = details[0].get("sem")
         

        slots = {item.get("slot") for item in details}
        for slot in slots:
            branches = [item.get("branch") for item in details if item.get("slot") == slot]
            students = []
            for branch in branches:
                file_path = f'./updatedExcels/Sem{sem}_{branch}.xlsx'
                
                try:
                    wb = openpyxl.load_workbook(file_path)
                    
                    sheetnames_lc_map = {s.lower(): s for s in wb.sheetnames} 
                    
                    slot_lc = slot.lower()
                    if slot_lc in sheetnames_lc_map:
                        ws_main = wb[sheetnames_lc_map[slot_lc]]
                        for row in ws_main.iter_rows(min_row=1, max_row=ws_main.max_row, values_only=True):
                            if any(row):
                                try:
                                    extracted = [
                                        row[0],  # Student_name
                                        row[1],  # Reg_No
                                        row[3]   # Sub_code
                                    ]
                                    if all(extracted):
                                        students.append(extracted + [branch])
                                except IndexError:
                                    continue
                            
                    
                except Exception as e:
                    print(f"Error opening {file_path}: {e}", file=sys.stderr)
                    sys.exit(1)
            total_students = len(students)
            assigned = 0
            room_data = []
            wb_output = openpyxl.Workbook()
        
            for room in rooms:
                if assigned >= total_students:
                    break

                room_name = room["room"]
                capacity = room["capacity"]
                ws = wb_output.create_sheet(title=room_name)
                ws.append(["Seat_no", "Student_name", "Reg_No", "Branch", "Sub_code"])
                
                row = 2  
                seat_no = 1
                for i in range(capacity):
                    if assigned >= total_students:
                        break
                    student = students[assigned]
                    seat_no = "{:03d}".format(i + 1)  
                    row_data = [seat_no] + student
                    for c, value in enumerate(row_data):
                        ws.cell(row=row, column=1 + c, value=value)  
                    row += 1
                    assigned += 1

                room_data.append({
                    "room": room_name,
                    "capacity": capacity,
                    "assigned": row - 2  
                })

            if "Sheet" in wb_output.sheetnames:
                del wb_output["Sheet"]
        if assigned < total_students:
                print(f"Error: Not enough capacity for {total_students} students.", file=sys.stderr)
                sys.exit(1)

        os.makedirs('./output', exist_ok=True)
        output_filename = f"./output/Seating_Sem{sem}_Slot_{slot}.xlsx"
        wb_output.save(output_filename)   
        
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()